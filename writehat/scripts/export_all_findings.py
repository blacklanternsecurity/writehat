import logging
import openpyxl
from unidecode import unidecode

import sys
from pathlib import Path
package_path = Path(__file__).resolve().parent.parent.parent
sys.path.append(str(package_path))

import os
os.environ['DJANGO_SETTINGS_MODULE'] = 'writehat.settings'
import django
django.setup()


from writehat.lib.finding import *
from writehat.lib.engagement import *
from writehat.lib.findingGroup import *

log = logging.getLogger(__name__)

def generateExcel(CVSSEngagementFindings, DREADEngagementFindings, ProactiveEngagementFindings, CVSSDatabaseFindings, DREADDatabaseFindings, ProactiveDatabaseFindings):

    workbook = openpyxl.Workbook()
    CVSSSheet = workbook.active
    DREADSheet = workbook.create_sheet()
    ProactiveSheet = workbook.create_sheet()
    CVSSSheet.title = "CVSS Engagement Findings"
    DREADSheet.title = "DREAD Engagement Findings"
    ProactiveSheet.title = "Proactive Engagement Findings"
    CVSSSheetDatabase = workbook.create_sheet()
    DREADSheetDatabase = workbook.create_sheet()
    ProactiveSheetDatabase = workbook.create_sheet()
    CVSSSheetDatabase.title = "CVSS Database Findings"
    DREADSheetDatabase.title = "DREAD Database Findings"
    ProactiveSheetDatabase.title = "Proactive Database Findings"


    ### CVSS Sheet ###

    # Define the titles for columns
    columns = ['name','category','background','description','affectedResources','proofOfConcept','remediation','toolsUsed','vector','severity','cvssScore','references','engagement']
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = CVSSSheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    
    # Iterate through CVSSEngagementFindings
    for i in CVSSEngagementFindings:
        log.debug("  Finding: {0}".format(i.id))
        row_num += 1

        # Define the data for each cell in row
        try:
            row = [
                str(i.name),
                ' -> '.join(i.category.getCategoryBreadcrumbs()[::-1]),
                str(i.background),
                str(i.description),
                str(i.affectedResources),
                str(i.proofOfConcept),
                str(i.remediation),
                str(i.toolsUsed),
                str(i.vector),
                str(i.cvss.severity),
                str(i.cvss.score),
                str(i.references),
                f'{i.parent.name} ({i.parent.id})',
            ]
        except Engagement.DoesNotExist:
            pass

        # Assign the data for each cell of the row

        # Assign the titles for each cell of the header
        for col_num, cell_value in enumerate(row, 1):
            cell = CVSSSheet.cell(row=row_num, column=col_num)
            try:
                cell.value = unidecode(cell_value)
            except:
                pass

    ### DREAD Sheet ###

    # Define the titles for columns
    columns = ['name', 'category', 'background', 'description', 'remediation', 'vector', 'severity', 'dreadScore', 'references', 'engagement']
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = DREADSheet.cell(row=row_num, column=col_num)
        cell.value = column_title

   # Iterate through DREADEngagementFindings
    for i in DREADEngagementFindings:
        log.debug("  Finding: {0}".format(i.id))
        row_num += 1

        try:
            row = [
                str(i.name),
                ' -> '.join(i.category.getCategoryBreadcrumbs()[::-1]),
                str(i.background),
                str(i.description),
                str(i.remediation),
                str(i.vector),
                str(i.dread.severity),
                str(i.dread.score),
                str(i.references),
                f'{i.parent.name} ({i.parent.id})',
            ]
        except Engagement.DoesNotExist:
            pass

        # Assign the titles for each cell of the header
        for col_num, cell_value in enumerate(row, 1):
            cell = DREADSheet.cell(row=row_num, column=col_num)
            try:
                cell.value = unidecode(cell_value)
            except:
                pass


    ### Proactive Sheet ###

    # Define the titles for columns
    columns = ['name', 'category', 'background', 'description', 'references', 'engagement']
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = ProactiveSheet.cell(row=row_num, column=col_num)
        cell.value = column_title

   # Iterate through ProactiveEngagementFindings
    for i in ProactiveEngagementFindings:
        log.debug("  Finding: {0}".format(i.id))
        row_num += 1


        try:
            row = [
                str(i.name),
                ' -> '.join(i.category.getCategoryBreadcrumbs()[::-1]),
                str(i.background),
                str(i.description),
                str(i.references),
                f'{i.parent.name} ({i.parent.id})',
            ]
        except Engagement.DoesNotExist:
            pass

    # Assign the titles for each cell of the header
        for col_num, cell_value in enumerate(row, 1):
            cell = ProactiveSheet.cell(row=row_num, column=col_num)
            try:
                cell.value = unidecode(cell_value)
            except:
                pass
    








    ### CVSS Sheet ###

    # Define the titles for columns
    columns = ['name','category','background','remediation','toolsUsed','vector','severity','cvssScore','references']
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = CVSSSheetDatabase.cell(row=row_num, column=col_num)
        cell.value = column_title
    
    # Iterate through CVSSEngagementFindings
    for i in CVSSDatabaseFindings:
        log.debug("  Finding: {0}".format(i.id))
        row_num += 1

        # Define the data for each cell in row
        row = [
            str(i.name),
            ' -> '.join(i.category.getCategoryBreadcrumbs()[::-1]),
            str(i.background),
            str(i.remediation),
            str(i.toolsUsed),
            str(i.vector),
            str(i.cvss.severity),
            str(i.cvss.score),
            str(i.references),
        ]

        # Assign the data for each cell of the row

        # Assign the titles for each cell of the header
        for col_num, cell_value in enumerate(row, 1):
            cell = CVSSSheetDatabase.cell(row=row_num, column=col_num)
            try:
                cell.value = unidecode(cell_value)
            except:
                pass

    ### DREAD Sheet ###

    # Define the titles for columns
    columns = ['name', 'category', 'background', 'remediation', 'vector', 'severity', 'dreadScore', 'references']
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = DREADSheetDatabase.cell(row=row_num, column=col_num)
        cell.value = column_title

   # Iterate through DREADEngagementFindings
    for i in DREADDatabaseFindings:
        log.debug("  Finding: {0}".format(i.id))
        row_num += 1

        row = [
            str(i.name),
            ' -> '.join(i.category.getCategoryBreadcrumbs()[::-1]),
            str(i.background),
            str(i.remediation),
            str(i.vector),
            str(i.dread.severity),
            str(i.dread.score),
            str(i.references),
        ]

        # Assign the titles for each cell of the header
        for col_num, cell_value in enumerate(row, 1):
            cell = DREADSheetDatabase.cell(row=row_num, column=col_num)
            try:
                cell.value = unidecode(cell_value)
            except:
                pass


    ### Proactive Sheet ###

    # Define the titles for columns
    columns = ['name', 'category', 'background', 'references']
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = ProactiveSheetDatabase.cell(row=row_num, column=col_num)
        cell.value = column_title

   # Iterate through ProactiveEngagementFindings
    for i in ProactiveDatabaseFindings:
        log.debug("  Finding: {0}".format(i.id))
        row_num += 1



        row = [
            str(i.name),
            ' -> '.join(i.category.getCategoryBreadcrumbs()[::-1]),
            str(i.background),
            str(i.references),
        ]

        # Assign the titles for each cell of the header
        for col_num, cell_value in enumerate(row, 1):
            cell = ProactiveSheetDatabase.cell(row=row_num, column=col_num)
            try:
                cell.value = unidecode(cell_value)
            except:
                pass


    return workbook


CVSSEngagementFindings = []
DREADEngagementFindings = []
ProactiveEngagementFindings = []
CVSSDatabaseFindings = []
DREADDatabaseFindings = []
ProactiveDatabaseFindings = []

for fgroup in BaseFindingGroup.filter_children():
    if fgroup.scoringType == "CVSS":
        CVSSEngagementFindings += fgroup.findings
    elif fgroup.scoringType == 'DREAD':
        DREADEngagementFindings += fgroup.findings
    elif fgroup.scoringType == 'PROACTIVE':
        ProactiveEngagementFindings += fgroup.findings
for finding in BaseDatabaseFinding.all_children():
    if finding.scoringType == "CVSS":
        CVSSDatabaseFindings.append(finding)
    elif finding.scoringType == 'DREAD':
        DREADDatabaseFindings.append(finding)
    elif finding.scoringType == 'PROACTIVE':
        ProactiveDatabaseFindings.append(finding)

workbook = generateExcel(
    CVSSEngagementFindings,
    DREADEngagementFindings,
    ProactiveEngagementFindings,
    CVSSDatabaseFindings,
    DREADDatabaseFindings,
    ProactiveDatabaseFindings
)
workbook.save('export.xlsx')
workbook.close()