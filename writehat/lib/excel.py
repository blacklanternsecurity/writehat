import logging
import openpyxl

log = logging.getLogger(__name__)

def generateExcel(CVSSEngagementFindings, DREADEngagementFindings, ProactiveEngagementFindings):

    workbook = openpyxl.Workbook()
    CVSSSheet = workbook.active
    DREADSheet = workbook.create_sheet()
    ProactiveSheet = workbook.create_sheet()
    CVSSSheet.title = "CVSS Findings"
    DREADSheet.title = "DREAD Findings"
    ProactiveSheet.title = "Proactive Findings"


    ### CVSS Sheet ###

    # Define the titles for columns
    columns = ['name','category','background','description','affectedResources','proofOfConcept','remediation','toolsUsed','vector','severity','cvssScore','references']
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = CVSSSheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    
    # Iterate through CVSSEngagementFindings
    for i in CVSSEngagementFindings:
        log.debug("  Finding: {0}".format(i.id))
        row_num += 1

        # Define the data for each cell in row
        row = [
            i.name,
            ' -> '.join(i.category.getCategoryBreadcrumbs()[::-1]),
            '"%s"' % i.background,
            '"%s"' % i.description,
            '"%s"' % i.affectedResources,
            '"%s"' % i.proofOfConcept,
            '"%s"' % i.remediation,
            '"%s"' % i.toolsUsed,
            i.vector,
            i.cvss.severity,
            i.cvss.score,
            '"%s"' % i.references,
        ]

        # Assign the data for each cell of the row

        # Assign the titles for each cell of the header
        for col_num, cell_value in enumerate(row, 1):
            cell = CVSSSheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    ### DREAD Sheet ###

    # Define the titles for columns
    columns = ['name', 'category', 'background', 'description', 'remediation', 'vector', 'severity', 'dreadScore', 'references']
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = DREADSheet.cell(row=row_num, column=col_num)
        cell.value = column_title

   # Iterate through DREADEngagementFindings
    for i in DREADEngagementFindings:
        log.debug("  Finding: {0}".format(i.id))
        row_num += 1

        row = [
            i.name,
            ' -> '.join(i.category.getCategoryBreadcrumbs()[::-1]),
            '"%s"' % i.background,
            '"%s"' % i.description,
            '"%s"' % i.remediation,
            i.vector,
            i.dread.severity,
            i.dread.score,
            '"%s"' % i.references,
        ]

        # Assign the titles for each cell of the header
        for col_num, cell_value in enumerate(row, 1):
            cell = DREADSheet.cell(row=row_num, column=col_num)
            cell.value = cell_value


    ### Proactive Sheet ###

    # Define the titles for columns
    columns = ['name', 'category', 'background', 'description', 'references']
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = ProactiveSheet.cell(row=row_num, column=col_num)
        cell.value = column_title

   # Iterate through ProactiveEngagementFindings
    for i in ProactiveEngagementFindings:
        log.debug("  Finding: {0}".format(i.id))
        row_num += 1


        row = [
            i.name,
            ' -> '.join(i.category.getCategoryBreadcrumbs()[::-1]),
            '"%s"' % i.background,
            '"%s"' % i.description,
            '"%s"' % i.references,
        ]

    # Assign the titles for each cell of the header
        for col_num, cell_value in enumerate(row, 1):
            cell = ProactiveSheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    
    return workbook

